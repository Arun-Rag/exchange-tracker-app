import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data=json.load(open('exchanges.json'))
BRAND="0B4F8A"; BRAND2="1769AA"; LIGHT="EAF2F8"
hdr_font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
cell_font=Font(name="Arial",size=10)
hdr_fill=PatternFill("solid",fgColor=BRAND)
zebra=PatternFill("solid",fgColor="F5F8FB")
open_fill=PatternFill("solid",fgColor="FEF3C7")
due_fill=PatternFill("solid",fgColor="FEE2E2")
closed_fill=PatternFill("solid",fgColor="DCFCE7")
thin=Side(style="thin",color="DDDDDD"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(vertical="top",wrap_text=True)

cols=[("ID",6),("Date",10),("Customer",15),("Sold P/N",14),("Sold S/N",14),("Sales Order No",13),
("Sales Price",12),("Currency",9),("Core P/N",14),("Core S/N",16),("Exchange Status",20),
("RO Number",12),("Repair Shop",13),("Assy Status",22),("Repair Quote Price",13),
("Failed Nozzles",9),("Ownership",16),("Comment",34),("Order Status",11)]
key=["ID","Date","Customer","SoldPN","SoldSN","SalesOrderNo","SalesPrice","SalesCurrency",
"CorePN","CoreSN","ExchangeStatus","RONumber","RepairShop","AssyStatus","RepairQuotePrice",
"FailedNozzles","Ownership","Comment","OrderStatus"]
# note: Ownership key is OwnershipChanged
key[16]="OwnershipChanged"

wb=Workbook()
ws=wb.active; ws.title="Exchanges"
ws.append([c[0] for c in cols])
for i,(name,w) in enumerate(cols,1):
    cell=ws.cell(1,i); cell.font=hdr_font; cell.fill=hdr_fill; cell.alignment=Alignment(vertical="center",wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width=w
for r,rec in enumerate(data,2):
    for i,k in enumerate(key,1):
        v=rec.get(k,"")
        if v is None: v=""
        ws.cell(r,i,v)
    # styling
    status=rec["OrderStatus"]; due=rec["Due"]
    for i in range(1,len(cols)+1):
        c=ws.cell(r,i); c.font=cell_font; c.border=border; c.alignment=wrap
        if r%2==0: c.fill=zebra
    # status cell color
    sc=ws.cell(r,len(cols))
    sc.fill=closed_fill if status=="Closed" else (due_fill if due else open_fill)
    sc.value = "Closed" if status=="Closed" else ("Due" if due else "Open")
    sc.font=Font(name="Arial",size=10,bold=True,
        color=("15803D" if status=="Closed" else ("B91C1C" if due else "B45309")))
    # price number format
    ws.cell(r,7).number_format='#,##0;(#,##0);-'
    ws.cell(r,15).number_format='#,##0;(#,##0);-'
ws.freeze_panes="A2"
ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{len(data)+1}"
ws.row_dimensions[1].height=30

# ---- Dashboard sheet ----
ds=wb.create_sheet("Dashboard",0)
ds.column_dimensions['A'].width=34; ds.column_dimensions['B'].width=16
ds['A1']="3rd Party Exchange Tracker — Summary"; ds['A1'].font=Font(name="Arial",bold=True,size=14,color=BRAND)
ds.append([])
n=len(data)+1
rng=f"Exchanges!S2:S{n}"
items=[("Total records", f'=COUNTA(Exchanges!A2:A{n})'),
 ("Open orders", f'=COUNTIF({rng},"Open")'),
 ("Due (core owed)", f'=COUNTIF({rng},"Due")'),
 ("Closed orders", f'=COUNTIF({rng},"Closed")'),
 ("Failed nozzles to claim", f'=SUM(Exchanges!P2:P{n})'),
 ("Total sales value (parsed, USD)", f'=SUM(Exchanges!G2:G{n})')]
r=3
for label,formula in items:
    ds.cell(r,1,label).font=Font(name="Arial",size=11,bold=True)
    c=ds.cell(r,2,formula); c.font=Font(name="Arial",size=11,color=BRAND2)
    ds.cell(r,1).fill=PatternFill("solid",fgColor=LIGHT)
    if "value" in label: c.number_format='#,##0'
    r+=1
ds.cell(r+1,1,"Open the Exchanges tab for the full list. Use the column filters to sort by Customer, Status, Repair Shop, etc.").font=Font(name="Arial",size=9,italic=True,color="6B7280")

wb.save("Exchanges_for_import.xlsx")
print("xlsx saved, rows:",len(data))
